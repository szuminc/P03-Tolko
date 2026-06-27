import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── helpers ──────────────────────────────────────────────────────────────────

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ── colour palette ────────────────────────────────────────────────────────────
PHASE_COLORS = {
    1: "1E3A5F",   # deep navy
    2: "2E6B4F",   # forest green
    3: "7B3F00",   # dark amber
    4: "4A235A",   # purple
    5: "1A5276",   # steel blue
    6: "784212",   # burnt orange
    7: "1B4F72",   # ocean
    8: "212121",   # near-black
}
PHASE_ACCENT = {
    1: "D6E4F0",
    2: "D5F5E3",
    3: "FDEBD0",
    4: "E8DAEF",
    5: "D6EAF8",
    6: "FDEBD0",
    7: "D6EAF8",
    8: "EAECEE",
}
TASK_BG = "F9F9F9"
HEADER_BG = "1C1C1C"

# ── Sheet 1: Phase Overview ───────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Phase Overview"

ws1.row_dimensions[1].height = 36
ws1.row_dimensions[2].height = 20
for c in range(1, 6):
    ws1.column_dimensions[get_column_letter(c)].width = [4, 14, 30, 38, 20][c - 1]

# Title row
ws1.merge_cells("A1:E1")
title_cell = ws1["A1"]
title_cell.value = "Workout PWA — Execution Plan"
title_cell.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
title_cell.fill = hex_fill(HEADER_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# Column headers
headers = ["#", "Phase Name", "Goal / Focus", "Output / Done-When", "Status"]
for col, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=col, value=h)
    cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    cell.fill = hex_fill("2C3E50")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()

phases_overview = [
    (1, "Project Setup",      "Running PWA shell installable on iPhone via Safari",
     "App installs to iPhone home screen and loads offline"),
    (2, "Movement Library",   "Data-driven movements with looping CSS animations",
     "All movements display with working animations on the library screen"),
    (3, "Routine Builder",    "Users can build, name, save, and manage routines",
     "User can create, edit, save, and delete routines"),
    (4, "Workout Player",     "Focused workout mode that guides user through a routine",
     "User can start a routine and be guided through every movement"),
    (5, "Milestones & Themes","Celebrate workout streaks with unlockable backgrounds",
     "Milestone-triggering workout shows a celebration and unlocks new background"),
    (6, "Daily Quote",        "Fresh motivational quote greets the user each day",
     "Home screen shows a daily quote that changes each day"),
    (7, "PWA Hardening",      "Reliable offline experience and smooth iOS installation",
     "App works fully offline and feels native on iPhone"),
    (8, "Polish & Testing",   "Smooth, enjoyable experience ready for daily use",
     "App is stable, smooth, and ready for daily use on iPhone"),
]

for row_i, (num, name, goal, done_when) in enumerate(phases_overview, 3):
    ws1.row_dimensions[row_i].height = 40
    accent = PHASE_ACCENT[num]
    row_data = [num, name, goal, done_when, "Not Started"]
    for col, val in enumerate(row_data, 1):
        cell = ws1.cell(row=row_i, column=col, value=val)
        cell.fill = hex_fill(accent)
        cell.font = Font(name="Calibri", size=10,
                         bold=(col in (1, 2)),
                         color=PHASE_COLORS[num] if col in (1, 2) else "333333")
        cell.alignment = Alignment(horizontal="center" if col in (1, 5) else "left",
                                   vertical="center", wrap_text=True)
        cell.border = border()

# ── Sheet 2: Task Tracker ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("Task Tracker")

col_widths = [4, 14, 52, 14, 22, 30]
col_letters = [get_column_letter(i) for i in range(1, 7)]
for ltr, w in zip(col_letters, col_widths):
    ws2.column_dimensions[ltr].width = w

# Title
ws2.merge_cells("A1:F1")
t = ws2["A1"]
t.value = "Workout PWA — Task Tracker"
t.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
t.fill = hex_fill(HEADER_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

# Column headers
task_headers = ["Ph", "Phase Name", "Task", "Status", "Priority", "Notes"]
for col, h in enumerate(task_headers, 1):
    cell = ws2.cell(row=2, column=col, value=h)
    cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    cell.fill = hex_fill("2C3E50")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border()
ws2.row_dimensions[2].height = 20

all_tasks = [
    (1, "Project Setup", [
        "Create GitHub repo",
        "Scaffold file structure (index.html, style.css, app.js, manifest.json, sw.js)",
        "Set up manifest.json with name, icons, display mode, theme colour",
        "Register a basic service worker (cache shell on install)",
        "Verify 'Add to Home Screen' works in Safari iOS",
        "Set up CSS custom properties for theming (colours, fonts, spacing)",
        "Build app shell layout: header, main content area, bottom nav",
    ]),
    (2, "Movement Library", [
        "Define movement data structure in data/movements.js",
        "Add initial movement set: Abs, Four-point rotation, Hip flexor, Plank hold + 4 more",
        "Write CSS keyframe animation for each movement in animations/movements.css",
        "Build a movement library screen listing all movements",
        "Tapping a movement shows detail view (name, description, animation, default settings)",
    ]),
    (3, "Routine Builder", [
        "Build routine list screen (home — shows saved routines)",
        "Build 'create routine' flow — name the routine",
        "Build 'create routine' flow — browse and select movements from library",
        "Build 'create routine' flow — set duration (seconds) and reps per movement",
        "Build 'create routine' flow — reorder movements (drag or up/down buttons)",
        "Build 'create routine' flow — remove movements",
        "Save routine to localStorage",
        "Edit and delete existing routines",
        "Validate: routine must have at least one movement before saving",
    ]),
    (4, "Workout Player", [
        "Build workout player screen",
        "Show current movement name, description, and animation (full-screen, minimal UI)",
        "Implement countdown timer per set (configurable duration)",
        "Show rep count and overall progress (e.g. 'Movement 2 of 5')",
        "Auto-advance to next movement when timer reaches zero",
        "Play audio cue on set end (Web Audio API — no external files)",
        "Pause / resume control",
        "Show completion screen when routine finishes",
        "Increment totalWorkoutsCompleted in localStorage on completion",
    ]),
    (5, "Milestones & Themes", [
        "Define milestone thresholds and corresponding background themes",
        "Check for new milestone on every workout completion",
        "Show milestone celebration screen (message + theme preview)",
        "Persist unlocked milestones and active background in localStorage",
        "Apply active background theme to app shell via CSS custom properties",
        "Build achievements view showing locked and unlocked milestones",
    ]),
    (6, "Daily Quote", [
        "Write a curated list of 60+ quotes in data/quotes.js",
        "Implement date-based quote selection (same quote all day, changes at midnight)",
        "Display quote prominently on home / routine list screen",
        "No external API — fully offline",
    ]),
    (7, "PWA Hardening", [
        "Implement cache-first service worker strategy for all app assets",
        "Pre-cache all HTML, CSS, JS, and animation files on install",
        "Handle service worker updates gracefully (notify user, refresh)",
        "Verify full offline functionality (no network required after first load)",
        "Add iOS-specific meta tags (apple-mobile-web-app-capable, status bar, splash screens)",
        "Test install and launch from iPhone home screen in Safari",
        "Prevent scroll bounce and ensure full-screen feel on iOS",
    ]),
    (8, "Polish & Testing", [
        "Test all flows on iPhone (Safari, installed PWA mode)",
        "Test on at least two iOS versions",
        "Accessibility: sufficient colour contrast, tap target sizes ≥ 44px",
        "Keyboard / focus management in modals and player",
        "Performance: animations run at 60fps, no jank during timer",
        "Edge cases: empty routine list, single-movement routine, back button during workout",
        "Add subtle transition animations between screens",
        "Final UX review: icon, splash screen, colour themes",
    ]),
]

current_row = 3
for phase_num, phase_name, tasks in all_tasks:
    accent = PHASE_ACCENT[phase_num]
    phase_color = PHASE_COLORS[phase_num]

    for task_i, task in enumerate(tasks):
        ws2.row_dimensions[current_row].height = 30
        row_vals = [phase_num, phase_name, task, "Not Started", "Medium", ""]
        for col, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=current_row, column=col, value=val)
            cell.fill = hex_fill(accent if col <= 2 else TASK_BG)
            is_phase_col = col in (1, 2)
            cell.font = Font(
                name="Calibri", size=10,
                bold=is_phase_col,
                color=phase_color if is_phase_col else "333333"
            )
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 4, 5) else "left",
                vertical="center",
                wrap_text=True
            )
            cell.border = border()
        current_row += 1

# ── Sheet 3: Tech Decisions ───────────────────────────────────────────────────
ws3 = wb.create_sheet("Tech Decisions")
for ltr, w in zip(["A", "B", "C"], [20, 22, 48]):
    ws3.column_dimensions[ltr].width = w

ws3.merge_cells("A1:C1")
t3 = ws3["A1"]
t3.value = "Key Technical Decisions"
t3.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
t3.fill = hex_fill(HEADER_BG)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

for col, h in enumerate(["Decision", "Choice", "Reason"], 1):
    cell = ws3.cell(row=2, column=col, value=h)
    cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    cell.fill = hex_fill("2C3E50")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border()
ws3.row_dimensions[2].height = 20

tech_rows = [
    ("Framework",  "None (vanilla JS)",       "Zero build tooling, easier PWA"),
    ("Animations", "CSS keyframes",            "Performant, no dependencies"),
    ("Storage",    "localStorage",             "Simple, offline, no backend"),
    ("Audio",      "Web Audio API",            "No external files, works offline"),
    ("Routing",    "Hash-based (#routines…)",  "No server needed"),
    ("Icons",      "Self-hosted PNG",          "No CDN dependency"),
]
row_colors = ["F2F3F4", "FFFFFF"]
for i, (dec, choice, reason) in enumerate(tech_rows, 3):
    ws3.row_dimensions[i].height = 28
    bg = row_colors[i % 2]
    for col, val in enumerate([dec, choice, reason], 1):
        cell = ws3.cell(row=i, column=col, value=val)
        cell.fill = hex_fill(bg)
        cell.font = Font(name="Calibri", size=10,
                         bold=(col == 1), color="1C1C1C")
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border()

# ── save ─────────────────────────────────────────────────────────────────────
out_path = r"C:\Users\SIMIN\ClaudeCode_szmn\Projects_szmn\P03-Tolko-Workout -PWA\doc\P03_Execution_Plan.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
